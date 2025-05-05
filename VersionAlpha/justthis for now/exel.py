import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import seaborn as sns

# Data
summary_data = {
    "Metric": [
        "Total samples", "Training samples", "Validation samples", "Test samples"
    ],
    "Value": [7861, "4716 (60.0%)", "1572 (20.0%)", "1573 (20.0%)"]
}

class_distribution = {
    "Class": ["No Fault", "LG", "LL BC", "LLG", "LLL", "LLLG"],
    "Samples": [2365, 1129, 1004, 1134, 1096, 1133]
}

feature_stats = {
    "Feature": ["Ia", "Ib", "Ic", "Va", "Vb", "Vc"],
    "Mean": [22.3695, -41.0013, 30.5312, -0.0076, -0.0021, 0.0098],
    "Std": [466.4818, 435.6345, 367.1496, 0.2908, 0.3172, 0.3080],
    "Min": [-883.5423, -900.5270, -883.3451, -0.6207, -0.6080, -0.6127],
    "25%": [-86.1375, -250.0775, -62.4135, -0.1382, -0.1781, -0.2177],
    "50%": [4.6081, 7.3404, -3.1538, -0.0027, -0.0015, 0.0123],
    "75%": [270.1579, 90.8189, 48.8416, 0.1142, 0.1519, 0.2477],
    "Max": [885.7386, 889.6850, 901.1012, 0.5953, 0.6279, 0.6002]
}

# Convert to DataFrames
summary_df = pd.DataFrame(summary_data)
class_df = pd.DataFrame(class_distribution)
features_df = pd.DataFrame(feature_stats)

# Create Excel workbook
wb = Workbook()
ws1 = wb.active
ws1.title = "Dataset Summary"

# Add summary
for r in dataframe_to_rows(summary_df, index=False, header=True):
    ws1.append(r)

# Add class distribution
ws2 = wb.create_sheet(title="Class Distribution")
for r in dataframe_to_rows(class_df, index=False, header=True):
    ws2.append(r)

# Add feature statistics
ws3 = wb.create_sheet(title="Feature Stats")
for r in dataframe_to_rows(features_df, index=False, header=True):
    ws3.append(r)

# Create and save charts
plt.figure(figsize=(6, 4))
sns.barplot(x="Class", y="Samples", data=class_df)
plt.title("Class Distribution")
plt.tight_layout()
plt.savefig("class_distribution.png")
plt.close()

plt.figure(figsize=(8, 6))
features_df.set_index("Feature")[["Mean", "Std"]].plot(kind="bar")
plt.title("Feature Mean and Std Dev")
plt.tight_layout()
plt.savefig("feature_stats.png")
plt.close()

# Insert charts into Excel
img1 = Image("class_distribution.png")
img2 = Image("feature_stats.png")
ws2.add_image(img1, "D2")
ws3.add_image(img2, "J2")

# Save workbook
excel_path = "fault_classification_summary.xlsx"
wb.save(excel_path)

excel_path
